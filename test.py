from PyPDF2 import PdfReader
import openpyxl

#demander nom du fichier pdf à récupérer
nom_fichier=input("Nom du fichier pdf à récupérer (sans le .pdf) : ")
#nom_fichier='test_analyse.pdf'

#demander nom du excel
#nom_excell=input("Nom du fichier excel : ")
#nom_excell='test_analyse.xlsx'

nom_excell=nom_fichier+".xlsx"

nom_fichier=nom_fichier+".pdf"

#demander si le fichier pdf est en anglais ou en français
langue=input("Langue du fichier pdf (fr ou en) : ")

if langue=="en":
    position_debut="Vulnerability"
    position_fin="10 Most"
else :
    position_debut="Aperçu"
    position_fin="10 fichiers"

wb = openpyxl.Workbook()
sheet = wb.active

reader = PdfReader(nom_fichier)
number_of_pages = len(reader.pages)

print("Nombre de pages du fichier pdf :",number_of_pages)

print("-------- 1 - trouves les bonnes pages ----------")

def find_pages(position_debut,position_fin):
    for i in range(number_of_pages):
        page = reader.pages[i]
        text = page.extract_text()
        if position_debut in text:
            page_debut=i
        if position_fin in text:
            page_fin=i
    return page_debut, page_fin

page_debut, page_fin = find_pages(position_debut,position_fin)
print("Page de début :",page_debut)
print("Page de fin :",page_fin)

nb_difference=page_fin-page_debut+1
print("Nombre de pages à récupérer :",nb_difference)


print("-------- 2 Extraction des pages ----------")


def extract_pages(page_debut,page_fin):
    page_variables = []
    textall = ""
    for i in range(page_debut,page_fin+1):
        #print("\n\nPage",i,":")
        global variable_name
        page = reader.pages[i]
        page_text = page.extract_text()
        variable_name = f"page_{i+1}_text"
        page_variables.append(variable_name)
        exec(f"{variable_name} = {repr(page_text)}")
        globals()[variable_name] = eval(variable_name)
        textall += page_text
    return page_variables, textall

page_variables, textall=extract_pages(page_debut,page_fin)

print("\n\nContenu de toutes les pages :", textall)

""""
print("\n\nListe des noms de variables :", page_variables)
for variable_name in page_variables:
    print(variable_name)
    if variable_name in globals():
        variable_content = eval(variable_name)
        print(f"Contenu de {variable_name}:")
        print(variable_content)
    else:
        print(f"La variable {variable_name} n'a pas été définie.")
"""

print("-------- 3 Extraction des données ----------")

def extract_data(textall):
    pos1 = textall.find(position_debut)
    pos2 = textall.find(position_fin)
    sousChaine = textall[pos1:pos2]
    return sousChaine

sousChaine=extract_data(textall)
print("Contenu de la sous chaine :", sousChaine)


print("-------- 4 Comptez nombre de ligne ----------")

# FONCTION RENVOYANT LE NOMBRE DE LIGNES D'UN FICHIER TEXTE
def countLigne(fichier):

    Liste=open(fichier,'r')
    i=1
    Ligne=Liste.readline()
    while Ligne!="":
        Ligne=Liste.readline()
        i+=1
    return i

print("-------- 5 Ecrire dans fichier ----------")

def ecris_fichier(sousChaine):
    fichier = open("data.txt", "w")
    fichier.write(sousChaine)
    fichier.close()

ecris_fichier(sousChaine)
nb_ligne=countLigne("data.txt")
print("Nombre de ligne :",nb_ligne)

nb_mot = len(sousChaine.split())
print("Nombre de mot :",nb_mot)


print("-------- 6 rendre jolie ----------")

def rendre_jolie(s):
    all_text2 = s.split("\n")
    print(f"all_text111: {all_text2}")
    all_text2 = [t for t in all_text2 if t.strip()] # Supprimer les éléments vides
    find_places = [i for i, text in enumerate(all_text2) if "PAGE" in text] # Trouver les positions des mots "PAGE"
    print(f"Les mots PAGE se trouvent aux positions : {find_places}")
    if not find_places: # S'il n'y a pas de mot "PAGE", retourner la liste d'origine
        return all_text2
    find_place = find_places[0] # Prendre la première occurrence
    string_garder = all_text2[find_place].split()
    print("string_garder :",string_garder)
    find_place2 = next((i for i, text in enumerate(string_garder) if "PAGE" in text), None) # Trouver la position du mot "PAGE" dans la sous-liste
    if find_place2 is not None:
        del string_garder[find_place2:find_place2+4] # Supprimer les éléments qui suivent le mot "PAGE"
    else: # Si le mot "PAGE" n'est pas trouvé dans la sous-liste, retourner la liste d'origine
        return all_text2
    all_text2[find_place] = " ".join(string_garder)
    print(f"\n\nall_text222: {all_text2}")
    return all_text2

print("-------- 7 boucle ----------")

def Recup(s, nb_ligne):
    all_text2 = rendre_jolie(s)
    h = 0
    test = 0
    nb_ligne = int(nb_ligne - 3)
    print("nombre de ligne:", nb_ligne)
    while h < nb_ligne:
        for row in all_text2:
            print(f"test: {test}")
            print(f"nb_ligne: {nb_ligne}")
            text_obtenu = list(row.split(" "))
            index_nombre = [i for i in range(len(text_obtenu)) if text_obtenu[i].isdigit()]
            if not index_nombre:
                # Si la liste est vide, passer à la ligne suivante
                continue
            index_nombre = index_nombre[0]
            print(f"index_nombre: {index_nombre}")
            avant_nombre, nombre, aprs_nombre = text_obtenu[:index_nombre], text_obtenu[index_nombre], text_obtenu[index_nombre+1:]
            #print(avant_nombre, nombre, aprs_nombre)
            c1 = sheet.cell(row=h+1, column=1)
            c1.value = ' '.join(avant_nombre)
            #print("c1.value :",c1.value)
            c2 = sheet.cell(row=h+1, column=2)
            c2.value = ''.join(nombre)
            #print("c2.value :",c2.value)
            c3 = sheet.cell(row=h+1, column=3)
            c3.value = ' '.join(aprs_nombre)
            #print("c3.value :",c3.value)
            h += 1
            test += 1
            print(f"test: {test}")
            print(f"nb_ligne: {nb_ligne}")
            if test == nb_ligne:
                break
    wb.save(nom_excell)

Recup(sousChaine,nb_ligne)