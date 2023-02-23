from PyPDF2 import PdfReader
import openpyxl

print("-------- 1 Paramétrages ----------")

#demander nom du fichier pdf à récupérer
nom_fichier=input("Nom du fichier pdf à récupérer (sans le .pdf) : ")

#pour test: 
#nom_fichier='test_analyse.pdf'

#demander nom du excel
#nom_excell=input("Nom du fichier excel : ")

#pour test :
#nom_excell='test_analyse.xlsx'

nom_excell=nom_fichier+".xlsx"

nom_fichier=nom_fichier+".pdf"

#demander si le fichier pdf est en anglais ou en français
langue=input("Langue du fichier pdf (fr ou en) : ")

if langue=="en":
    position_debut="Vulnerability"
    position_fin="10 Most"
else :
    position_debut="Type"
    position_fin="10 fichiers"

wb = openpyxl.Workbook()
sheet = wb.active

reader = PdfReader(nom_fichier)
number_of_pages = len(reader.pages)

print("Nombre de pages du fichier pdf :",number_of_pages)

print("-------- 2 trouves les bonnes pages ----------")

def find_pages(position_debut,position_fin):
    for i in range(number_of_pages):
        page = reader.pages[i]
        text = page.extract_text()
        if position_debut in text:
            page_debut=i
        if position_fin in text:
            page_fin=i
    return page_debut, page_fin

page_debut, page_fin = find_pages(position_debut,position_fin)
print("Page de début :",page_debut)
print("Page de fin :",page_fin)

nb_difference=page_fin-page_debut+1
print("Nombre de pages à récupérer :",nb_difference)


print("-------- 3 Extraction des pages ----------")


def extract_pages(page_debut,page_fin):
    page_variables = []
    textall = ""
    for i in range(page_debut,page_fin+1):
        #print("\n\nPage",i,":")
        global variable_name
        page = reader.pages[i]
        page_text = page.extract_text()
        variable_name = f"page_{i+1}_text"
        page_variables.append(variable_name)
        exec(f"{variable_name} = {repr(page_text)}")
        globals()[variable_name] = eval(variable_name)
        textall += page_text
    return page_variables, textall

page_variables, textall=extract_pages(page_debut,page_fin)

print("Contenu de toutes les pages :", textall)

""""
print("\n\nListe des noms de variables :", page_variables)
for variable_name in page_variables:
    print(variable_name)
    if variable_name in globals():
        variable_content = eval(variable_name)
        print(f"Contenu de {variable_name}:")
        print(variable_content)
    else:
        print(f"La variable {variable_name} n'a pas été définie.")
"""

print("-------- 4 Extraction des données ----------")

def extract_data(textall):
    pos1 = textall.find(position_debut)
    pos2 = textall.find(position_fin)
    sousChaine = textall[pos1:pos2]
    return sousChaine

sousChaine=extract_data(textall)
print("Contenu de la sous chaine :", sousChaine)


print("-------- 5 Comptez nombre de ligne ----------")

# FONCTION RENVOYANT LE NOMBRE DE LIGNES D'UN FICHIER TEXTE
def countLigne(fichier):

    Liste=open(fichier,'r')
    i=1
    Ligne=Liste.readline()
    while Ligne!="":
        Ligne=Liste.readline()
        i+=1
    return i

print("Nombre de ligne1 :",sousChaine.count("\n")+1)

print("-------- 6 Ecrire dans fichier ----------")

def ecris_fichier(sousChaine, name):
    fichier = open(name, "w")
    fichier.write(sousChaine)
    fichier.close()

ecris_fichier(sousChaine, "data.txt")
nb_ligne=countLigne("data.txt")
print("Nombre de lignes dans fichier :",nb_ligne)

nb_mot = len(sousChaine.split())
print("Nombre de mot :",nb_mot)


print("-------- 7 rendre jolie ----------")
"""def rendre_jolie(s):
    all_text2 = s.split("\n")
    print(f"all_text1: {all_text2}")
    del all_text2[-1]
    del all_text2[0]
    for i, text in enumerate(all_text2):
        if "PAGE" in text:
            print("Le mot PAGE se trouve à la position :", i)
            find_place=i
            break
    print("place :",find_place)
    string_garder=all_text2[find_place]
    print("string_garder :",string_garder)
    string_garder = string_garder.split()
    print("string_garder :",string_garder)
    for i, text in enumerate(string_garder):
        if "PAGE" in text:
            print("Le mot PAGE se trouve à la position :", i)
            find_place2=i
            break
    del string_garder[find_place2:find_place2+4]
    print("string_garder :",string_garder)
    del all_text2[find_place]
    all_text2.insert(find_place, " ".join(string_garder))
    print(f"\n\nall_text222: {all_text2}")
    return all_text2
    """

print("-------- 8 boucle ----------")

def Recup(s, nb_ligne):
    #all_text2 = rendre_jolie(s)
    all_text2 = s.split("\n")
    del all_text2[-1]
    del all_text2[0]
    print(f"all_text2: {all_text2}")
    chaine="\n".join(all_text2)
    print("\nchaine :",chaine)
    chaine.count("\n")
    ecris_fichier(chaine, "data2.txt")
    h = 0
    nb_ligne = chaine.count("\n")+1
    print("nombre de ligne:", nb_ligne)
    for row in all_text2:
        text_obtenu = list(row.split(" "))
        #print(f"text_obtenu: {text_obtenu}")
        index_nombre = [i for i in range(len(text_obtenu)) if text_obtenu[i].isdigit()]
        if not index_nombre:
            # Si la liste est vide, passer à la ligne suivante
            print("\n\nWTF Pas de nombre dans cette ligne", h, ":", text_obtenu)
            continue
        index_nombre = index_nombre[0]
        #print(f"index_nombre: {index_nombre}")
        avant_nombre, nombre, aprs_nombre = text_obtenu[:index_nombre], text_obtenu[index_nombre], text_obtenu[index_nombre+1:]
        #print(avant_nombre, nombre, aprs_nombre)
        c1 = sheet.cell(row=h+1, column=1)
        c1.value = ' '.join(avant_nombre)
        #print("c1.value :",c1.value)
        c2 = sheet.cell(row=h+1, column=2)
        c2.value = ''.join(nombre)
        #print("c2.value :",c2.value)
        c3 = sheet.cell(row=h+1, column=3)
        c3.value = ' '.join(aprs_nombre)
        #print("c3.value :",c3.value)
        h += 1
        print(f"h: {h}")
        if h == nb_ligne:
            print("Fin de boucle")
            break
    wb.save(nom_excell)

Recup(sousChaine,nb_ligne)