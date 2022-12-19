from PyPDF2 import PdfReader
import openpyxl

#demander nom du excel
#nom_excell=input("Nom du fichier excel : ")
nom_excell="test.xlsx"

"""
#demander position
position1=input("Position 1 : ")
#position2=input("Position 2 : ")
position3=input("Position 3 : ")
position4=input("Position 4 : ")
position5=input("Position 5 : ")"""

position1='Reflected'
#position2=input("Position 2 : ")
position3='Client'
position4='Stored'
position5='Open'

#demander nom du fichier pdf
#nom_fichier=input("Nom du fichier pdf : ")
nom_fichier='02789_gpf_dsifs_checkmarx_20221130-154117.pdf'

wb = openpyxl.Workbook()
sheet = wb.active

reader = PdfReader(nom_fichier)
number_of_pages = len(reader.pages)

print("-------- 1 Extraction des pages ----------")

page = reader.pages[11]
text = page.extract_text()
print("Page 1 :",text) 

page2 = reader.pages[12]
text2 = page2.extract_text()
print("Page 2 :",text2) 

page3 = reader.pages[13]
text3 = page3.extract_text()
print("Page 3 :",text3) 

page4 = reader.pages[14]
text4 = page4.extract_text()
print("Page 4 :",text4) 

print("-------- 2 Sortie des données page 12 ----------")

pos1 = text.find(position1)
pos2 = text.find('Z')
sousChaine = text[pos1:pos2]
print (sousChaine)

print("-------- 3 Sortie des données page 13 ----------")

pos3 = text2.find(position3)
pos4 = text2.find('Z')
sousChaine2 = text2[pos3:pos4]
print (sousChaine2)

print("-------- 3 Sortie des données page 14 ----------")

pos5 = text3.find(position4)
pos6 = text3.find('Z')
sousChaine3 = text3[pos5:pos6]
print (sousChaine3)

print("-------- 3 Sortie des données page 15 ----------")

pos7 = text4.find(position5)
pos8 = text4.find('10 Most')
sousChaine4 = text4[pos7:pos8]
print (sousChaine4)

print("-------- 4 Concaténation données ----------")

pdf_total=sousChaine+"\n"+sousChaine2+"\n"+sousChaine3+"\n"+sousChaine4
print(pdf_total)
pdf_total.splitlines()

print("-------- 5 Comptez nombre de ligne ----------")

# FONCTION RENVOYANT LE NOMBRE DE LIGNES D'UN FICHIER TEXTE
def countLigne(fichier):

    Liste=open(fichier,'r')
    i=1
    Ligne=Liste.readline()
    while Ligne!="":
        Ligne=Liste.readline()
        i+=1
    return i

print("-------- 6 Ecrire dans fichier ----------")
"""
fichier = open("data.txt", "w")
fichier.write(pdf_total)
fichier.close()"""
nb_ligne=countLigne("data.txt")
print("Nombre de ligne :",nb_ligne)
nb_mot = len(pdf_total.split())
print("Nombre de mot :",nb_mot)

fichier = open("data.txt", "r")
pdf_total=fichier.read()
fichier.close()
print(pdf_total)

print("-------- 7 boucle ----------")

def Recup(s,nb_ligne):
    all_text2 = s.split("\n")
    del all_text2[-1]
    print(f"all_text: {all_text2}")
    h=0
    test=0
    nb_ligne=int(nb_ligne)
    print("nombre de ligne:",nb_ligne)
    for row in all_text2:
            print("ROW=",row)
            #print(f"test1: {test}")
            print(f"nb_ligne1: {nb_ligne}")
            text_obtenu = list(row.split(" "))
            print('Textobentu=',text_obtenu)
            index_nombre = [i for i in range(len(text_obtenu)) if text_obtenu[i].isdigit()][0]
            print(f"index_nombre: {index_nombre}")
            avant_nombre, nombre, aprs_nombre = text_obtenu[:index_nombre], text_obtenu[index_nombre], text_obtenu[index_nombre+1:]
            #print(avant_nombre, nombre, aprs_nombre)
            c1=sheet.cell(row=h+1, column=1)
            c1.value=' '.join(avant_nombre)
            #print("c1.value :",c1.value)
            c2=sheet.cell(row=h+1, column=2)
            c2.value=''.join(nombre)
            #print("c2.value :",c2.value)
            c3=sheet.cell(row=h+1, column=3)
            c3.value=' '.join(aprs_nombre)
            #print("c3.value :",c3.value)
            h+=1
            test=test+1
            print(f"test2: {test}")
            print(f"H: {h}")
            print(f"nb_ligne2: {nb_ligne}")
            if test==nb_ligne:
                break
            if int(test)==50:
                print("wwhat")
                wb.save(nom_excell)
    wb.save(nom_excell)
        

Recup(pdf_total,nb_ligne)


